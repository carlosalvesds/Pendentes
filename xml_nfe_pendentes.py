import streamlit as st
import pandas as pd
import zipfile
import os
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="📥 Extrator XML NF-e", layout="wide")
st.title("📥 Extrator de Dados de XML | NF-e")

uploaded_zip = st.file_uploader("Envie um arquivo .zip contendo os XMLs (pode ter subpastas)", type=["zip"])

def extrair_dados_xml(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Adiciona namespace para evitar problemas
        ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}

        nNF = root.findtext('.//ns:ide/ns:nNF', default='', namespaces=ns)
        dhEmi_raw = root.findtext('.//ns:ide/ns:dhEmi', default='', namespaces=ns)
        dhEmi = ''
        if dhEmi_raw:
            try:
                dhEmi = datetime.fromisoformat(dhEmi_raw.replace('Z', '')).strftime('%d/%m/%Y %H:%M')
            except Exception:
                dhEmi = dhEmi_raw

        cnpj_emit = root.findtext('.//ns:emit/ns:CNPJ', default='', namespaces=ns)
        nome_emit = root.findtext('.//ns:emit/ns:xNome', default='', namespaces=ns)
        vNF = root.findtext('.//ns:ICMSTot/ns:vNF', default='', namespaces=ns)
        xMotivo = root.findtext('.//ns:protNFe/ns:infProt/ns:xMotivo', default='', namespaces=ns)

        # Produtos
        produtos = [prod.text for prod in root.findall('.//ns:det/ns:prod/ns:xProd', namespaces=ns)]
        mais_de_tres = len(produtos) > 3
        produtos_limitados = produtos[:3]
        produtos_join = " / ".join(produtos_limitados) + (" / ..." if mais_de_tres else "")



        return {
            "Número NF": nNF,
            "CNPJ Emitente": cnpj_emit,
            "Nome Emitente": nome_emit,
            "Data Emissão": dhEmi,
            "Valor NF": float(vNF) if vNF else 0.0,
            "Status NF": xMotivo,
            "Produtos": produtos_join,
            "Observações": ""
        }
    except Exception as e:
        return None

def aplicar_formatacao_excel(writer, sheet_name):
    ws = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value is not None else "") for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Se a coluna for "Valor NF", aplica formatação de moeda
            if cell.column_letter == 'E':  # Coluna E = "Valor NF"
                cell.number_format = 'R$ #,##0.00'


if uploaded_zip:
    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = os.path.join(tmpdir, "temp.zip")
        with open(zip_path, "wb") as f:
            f.write(uploaded_zip.read())

        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(tmpdir)

        xml_files = []
        for root_dir, _, files in os.walk(tmpdir):
            for file in files:
                if file.lower().endswith(".xml"):
                    xml_files.append(os.path.join(root_dir, file))

        st.success(f"{len(xml_files)} arquivos XML encontrados!")

        dados_extraidos = []
        for xml_file in xml_files:
            dados = extrair_dados_xml(xml_file)
            if dados:
                dados_extraidos.append(dados)

        if dados_extraidos:
            df = pd.DataFrame(dados_extraidos)

            # Converte "Data Emissão" para datetime para permitir ordenação
            df['Data Emissão'] = pd.to_datetime(df['Data Emissão'], format="%d/%m/%Y %H:%M", errors='coerce')

            # Ordena pela data (do mais antigo para o mais recente)
            df = df.sort_values(by="Data Emissão")

            # Formata para exibir apenas a data (sem hora)
            df['Data Emissão'] = df['Data Emissão'].dt.strftime('%d/%m/%Y')


            # Exportar para Excel com formatação
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Notas_Fiscais")
                aplicar_formatacao_excel(writer, "Notas_Fiscais")

            st.download_button(
                label="📥 Baixar Excel",
                data=output.getvalue(),
                file_name="dados_extraidos_nfe.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Nenhum dado foi extraído. Verifique os arquivos XML.")
